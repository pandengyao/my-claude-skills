#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
皮肤病多模态标注数据 Excel → JSON 转换脚本
用法: python excel_to_annotation_json.py input.xlsx output.json [--sheet 有效数据]
"""

import json
import sys
import argparse
from openpyxl import load_workbook


# Excel 列名 → JSON annotation_list 中 latitude_name 的映射
ANNOTATION_COLUMNS = [
    ("疾病词1【多选】", "疾病词1【多选】"),
    ("疾病词2【多选】", "疾病词2【多选】"),
    ("疾病词3【多选】", "疾病词3【多选】"),
    ("皮肤损害",        "皮肤损害【多选】"),
    ("最大皮损颜色",    "最大皮损颜色"),
    ("最大皮损大小",    "最大皮损大小"),
    ("皮损排列（多选）","皮损排列（多选）"),
    ("皮损数量",        "皮损数量"),
    ("皮损边缘",        "皮损边缘"),
    ("皮损界限",        "皮损界限"),
    ("表面湿度",        "表面湿度"),
    ("鳞屑/痂",         "鳞屑/痂"),
    ("年龄",            "年龄"),
    ("部位",            "部位"),
]


def parse_tags(value):
    """将 Excel 单元格值解析为 tags_list 数组"""
    if value is None or str(value).strip() == "":
        return []
    return [t.strip() for t in str(value).split(",") if t.strip()]


def parse_coordinates(value):
    """将 '4,17,584,319' 解析为 [4, 17, 584, 319]"""
    if value is None or str(value).strip() == "":
        return []
    try:
        return [int(x.strip()) for x in str(value).split(",")]
    except ValueError:
        return str(value).split(",")


def build_annotation(excel_value, latitude_name):
    """构建单个 annotation 项（不输出 label 字段）"""
    tags = parse_tags(excel_value)
    return {
        "latitude": {
            "latitude_name": latitude_name,
            "tags_list": tags,
        }
    }


def convert_row(row_dict):
    """将一行 Excel 数据转换为目标 JSON 格式的一条记录"""
    talk_id = str(row_dict.get("talk_id", "") or "")
    query = str(row_dict.get("query", "") or "")
    history = str(row_dict.get("history", "") or "")
    image_url = str(row_dict.get("image_url", "") or "")
    box_url = str(row_dict.get("box_url", "") or "")
    img_coordinates = parse_coordinates(row_dict.get("img_coordinates"))

    annotation_list = []
    for excel_col, json_name in ANNOTATION_COLUMNS:
        val = row_dict.get(excel_col)
        annotation_list.append(build_annotation(val, json_name))

    record = {
        "talk_id": [talk_id] if talk_id else [""],
        "query": [query] if query else [""],
        "history": [history] if history else [""],
        "image_url": [image_url] if image_url else [""],
        "box_url": [box_url] if box_url else [""],
        "img_coordinates": img_coordinates,
        "annotation_list": annotation_list,
    }
    return record


def convert_excel_to_json(input_path, output_path, sheet_name=None):
    """主转换函数"""
    wb = load_workbook(input_path, data_only=True)

    if sheet_name:
        ws = wb[sheet_name]
    else:
        if "有效数据" in wb.sheetnames:
            ws = wb["有效数据"]
        else:
            ws = wb.active

    # 读取表头（仅读有值的列）
    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val is None:
            break
        headers.append(val)

    num_cols = len(headers)
    print(f"检测到 {num_cols} 个有效列: {headers}")

    # 逐行转换
    result = []
    row_idx = 2
    while True:
        first_cell = ws.cell(row=row_idx, column=1).value
        if first_cell is None:
            break

        row_dict = {}
        for col in range(num_cols):
            row_dict[headers[col]] = ws.cell(row=row_idx, column=col + 1).value
        
        record = convert_row(row_dict)
        result.append([record])
        row_idx += 1

    total = len(result)
    print(f"共转换 {total} 条数据")

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"输出文件: {output_path}")
    return total


def main():
    parser = argparse.ArgumentParser(description="皮肤病标注数据 Excel → JSON 转换")
    parser.add_argument("input", help="输入 Excel 文件路径")
    parser.add_argument("output", help="输出 JSON 文件路径")
    parser.add_argument("--sheet", default=None, help="指定 sheet 名称（默认自动选择）")
    args = parser.parse_args()

    convert_excel_to_json(args.input, args.output, args.sheet)


if __name__ == "__main__":
    main()
